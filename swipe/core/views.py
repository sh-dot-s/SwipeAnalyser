import calendar
import os
import threading
from time import sleep

import nested_dict
# import numpy as np
# import pandas as pd
import openpyxl

from django.conf import settings
from django.contrib import admin
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.forms import (SetPasswordForm, UserChangeForm, UserCreationForm)
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import Permission, User, Group
from django.contrib.contenttypes.models import ContentType
from django.core import serializers
from django.core.files.storage import FileSystemStorage
from django.db import models, transaction
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.csrf import (csrf_exempt, csrf_protect, requires_csrf_token)
from django_filters.views import FilterView
from django_tables2 import MultiTableMixin, RequestConfig, SingleTableView
from django_tables2.export.views import ExportMixin

from swipe.core.corelogic import *
from swipe.core.filters import DocumentFilter, EmployeeFilter
from swipe.core.forms import *
from swipe.core.mail import *
from swipe.core.models import Document, Employee, EmployeeComplete
from swipe.core.tables import *

quoteFile = open(os.path.join(settings.STATICFILES_DIRS[0],"quotes.json"),'r')
quotes = json.load(quoteFile)
result = None
create_perm = Permission.objects.filter(codename='can_create_or_remove')

def logi(request):
    flag = None
    username = request.POST['username']
    password = request.POST['password']
    redir = (request.META['QUERY_STRING'])
    userExists = User.objects.filter(username=username)
    print(userExists[0].is_active)
    if len(userExists) > 0:
        flag = True
    else:
        flag = False
    user = authenticate(username=username, password=password)
    if user is not None:
        # request.session.set_expiry(1000)
        login(request, user=user)
        if not user.is_superuser:
            admin.site.unregister(User)
            admin.site.unregister(Group)
        else:
            try:
                admin.site.register(User)
                admin.site.register(Group)
            except:
                pass
        try:
            return redirect('{}'.format(redir.replace("next=","")))
        except:
            return redirect('/')            
    else:
        return render(request, 'registration/login.html', {'userExists': flag, 'state': userExists[0].is_active})

def logout_view(request):
    logout(request)
    return render(request, 'registration/logout.html')

@login_required
def home(request):
    documents = Document.objects.all()
    return render(request, 'home.html', {
        'documents': documents
    })


@csrf_exempt
@login_required
def upload(request):
    if request.method == 'POST' and request.FILES['input-b7[]']:
        fileXL = request.FILES['input-b7[]']
        print(fileXL)
        try:
            if fileXL.name in [documents.file_title for documents in Document.objects.all()]:
                return JsonResponse({'error': '%sFile already exists'%fileXL.name})
        except Exception as e:
            print(e)
            return JsonResponse({'error': '%s upload failed'%fileXL.name})
        else:
            fs = FileSystemStorage()
            filename = fs.save(fileXL.name, fileXL)
            uploaded_file_url = str(fs.url(filename))
            doc = Document(upload_url = uploaded_file_url, file_title = fileXL.name)
            doc.save()

            documents = Document.objects.all()
            return JsonResponse({})
    unProcessed = Document.objects.filter(processed=False)
    Docs = False
    if not len(unProcessed) > 0:
        Docs = True
    return render(request, 'upload.html', {"nogo": unProcessed, "nodocs" : Docs})

def deleteFiles(request):
    if request.method == 'POST':
        fileList = request.POST.getlist("check[]")
        for files in fileList:
            os.remove(os.path.join(settings.MEDIA_ROOT, files))
            Document.objects.filter(file_title=files).delete()
        return JsonResponse({"response": "success"})
    return HttpResponse('Unauthorized', status=401)

class FilteredDocument(LoginRequiredMixin, FilterView, SingleTableView):
    table_class = EmployeeTable
    model = Employee
    template_name = 'filter.html'
    filterset_class = EmployeeFilter
    table_pagination = {
        'per_page': 500
    }

@login_required
def analys(request):
    table = BootstrapTable(Employee.objects.all()) #order_by='-employee_id'
    RequestConfig(request, paginate={'per_page': 50}).configure(table)
    # print(RequestConfig)
    for objs in request:
        print(request)
    return render(request, 'analyse.html',{'documents': table})

@permission_required('auth.can_create_or_remove', raise_exception=True)
def create(request):
    users = User.objects.all()
    table = UserTable(users)
    RequestConfig(request, paginate={'per_page': 10}).configure(table)
    createForm = CustomUserCreationForm()
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            print('user created')
            users = User.objects.all()
            table = UserTable(users)
            RequestConfig(request, paginate={'per_page': 10}).configure(table)
            # users.refresh_from_db()
        else:
            # print(form.errors)
            return render(request, 'create_delete.html', {'users': table, 'count':len(users), 'form': form})
        # return render(request, 'userchange.html', {'form': form})
    if request.is_ajax():
        if request.GET.get('id') == 'formbtn-userchange':
            user = get_object_or_404(User, pk=request.GET.get('key'))
            form = CustomUserChangeForm(request.POST or None, instance=user)
            return render(request, 'userchange.html', {'form': form})

        if request.GET.get('id') == 'formbtn-changepwd':
            user = get_object_or_404(User, pk=request.GET.get('key'))
            form = SetPasswordForm(user)
            return render(request, 'userchange.html', {'form': form})

        if request.GET.get('id') == 'confirm':
            user = get_object_or_404(User, pk=request.GET.get('key'))
            user.delete()
            return HttpResponse("<h3>The user has been deleted succesfully... Reloading...... </h3>")
    return render(request, 'create_delete.html', {'users': table, 'count':len(User.objects.filter(is_superuser=False)), 'form': createForm})

@csrf_exempt
@login_required
def test(request):
    global result
    try:
        email = str(request.user.email)
        t = threading.Thread(target=start_processing, args=(email,), kwargs={})
        t.setDaemon(True)
        status = None
        if request.method == 'GET':
            if request.is_ajax():
                if len(result) > 0:
                    return JsonResponse({'missing': result, 'status':'stopped'})
                if t.isAlive():
                    return JsonResponse({'status': 'running'})
                else:
                    try:
                        for processFile in Document.objects.all():
                            if not processFile.processed and not status:
                                status = True
                        if not status:
                            return JsonResponse({'status': 'completed'})
                            result = None
                        else:
                            return JsonResponse({'status': 'running'})
                    except Exception as e:
                        print(e)
                        return JsonResponse({'status': 'running'})
            else:
                unProcessed = Document.objects.filter(processed=False)
                Docs = False
                if not len(unProcessed) > 0:
                    Docs = True
                return render(request, 'test.html', {"nogo": unProcessed, "nodocs" : Docs})
        else:
            result = validateData(request)
            if len(result) > 0:
                print(result)
                return JsonResponse({'missing': result, 'status':'stopped'})
            else:
                t.start()
                return JsonResponse({'status': 'started'})
    except Exception as e:
        return JsonResponse({'status': e})

@transaction.atomic
def start_processing(user):
    print("thread started")
    completeFiles = []
    for processFile in Document.objects.all():
            if not processFile.processed:
                masterEmployee, employeeComplete = load_master_emp(processFile.file_title)
                for cell in employeeComplete:
                    dateEmp, timeEmp = cell[0].value,cell[1].value
                    site, card, empid, cid = int(cell[2].value),int(cell[3].value),str(cell[4].value),int(cell[8].value)
                    empname, dept, typ, gate, inout, remark = str(cell[5].value),str(cell[6].value),str(cell[7].value),str(cell[9].value),str(cell[10].value),str(cell[11].value)
                    try:
                        if isinstance(dateEmp, datetime.datetime):
                            formattedDate = dateEmp.date()
                        else:
                            formattedDate = datetime.datetime.strptime(dateEmp, "%d/%m/%Y").date()
                    except:
                        if isinstance(dateEmp, datetime.datetime):
                            formattedDate = dateEmp.date()
                        else:
                            formattedDate = datetime.datetime.strptime(dateEmp, "%d-%m-%Y").date()

                    if isinstance(timeEmp, datetime.datetime):
                        formattedTime = timeEmp.time()
                    else:
                        formattedTime = (datetime.datetime.strptime(timeEmp, "%H:%M:%S")).time()

                    q=EmployeeComplete(dat=formattedDate,tim=formattedTime,sitecode=site,cardid=card,empid=empid,empname=empname,department=dept,typ=typ,cid=cid,gate=gate,inout=inout,remark=remark, document= processFile)
                    q.save()

                shiftDict,employeeShift,irregularShifts = shift_emp_masters()
                timeDict, d1, d2 = calc_time(masterEmployee,employeeShift,irregularShifts,shiftDict,processFile.file_title)
                delta = d2 - d1
                datesRange = []
                for i in range(delta.days + 1):
                    d = str(d1 + datetime.timedelta(days=i))
                    datesRange.append(d)
                for keys in timeDict.keys():
                    for dates in datesRange:
                        if not Employee.objects.filter(employee_id=keys,attendence_date=dates).exists():
                            day = calendar.day_name[datetime.datetime.strptime(dates, "%Y-%m-%d").weekday()]
                            q = Employee(employee_id=keys,attendence_date=dates,day=day,work_time="NS",employee_name=employeeShift[keys][2],dept=employeeShift[keys][1],tower=employeeShift[keys][0],code=employeeShift[keys][-1], document= processFile)
                            q.save()
                for keys,value in timeDict.items_flat():
                    if Employee.objects.filter(employee_id=keys[0],attendence_date=keys[1]).exists():
                        empObj = Employee.objects.filter(employee_id=keys[0],attendence_date=keys[1]).update(work_time=value)
                processFile.file_from = datesRange[0]
                processFile.file_to = datesRange[-1]
                processFile.processed = True
                processFile.records = len(employeeComplete)
                processFile.save(update_fields=["processed","file_to","file_from",'records'])
                completeFiles.append(processFile.file_title)
                print("process complete")
    sendCompleteMail(completeFiles, user)
    return

def table(request):
    return render(request, 'email.html')

@login_required
@csrf_exempt
def mail(request):
    table =  None
    if request.method == "POST":
        pks = request.POST.getlist("empid")
        selected_objects = Employee.objects.filter(pk__in=pks)
        table = BootstrapTable(selected_objects) #order_by='-employee_id'
        RequestConfig(request, paginate={'per_page': 10}).configure(table)
        sendSMTPMail(request, selected_objects)
    return render(request, 'selection.html', {'table': table})

@csrf_exempt
@login_required
def mail_response(request):
    if request.method == 'POST':
        print(request.POST)
        for keys in request.POST:
            response = (request.POST.getlist(keys))
            details = keys.split("_")
            date = datetime.datetime.strptime(details[2],"%Y-%m-%d").date()
            response = response[0]
            if 'comp' not in details[0]:
                Employee.objects.filter(employee_id=details[1], attendence_date=date).update(emp_response = response)
            else:
                # print(response)
                if response == "":
                    pass
                else:
                    response = "Comp off against "+response
                    Employee.objects.filter(employee_id=details[1], attendence_date=date).update(emp_response = response)
             # Review(response = response, employee_id = details[1], attendence_date=date)#, file_from=request.POST["from"], file_to=request.POST["to"]
        return HttpResponse("<h4>Response recorded, Thank you for your time.</h4>")
    return HttpResponse("<h3>Invalid Session</h3>")

@login_required
def details(request, item_pk):
    item_pk = item_pk.split("&")
    # print(item_pk)
    date = datetime.datetime.strptime(item_pk[1], "%Y-%m-%d").date()
    table = EmployeeComplete.objects.filter(
        empid=item_pk[0],
        dat__year=date.year,
        dat__month=date.month,
        dat__day=date.day
    )
    response = serializers.serialize('json', table)
    return HttpResponse(response, content_type='application/json')

@login_required
def ifDetailsPresent(request, item_pk):
    item_pk = item_pk.split("&")
    present = False
    date = datetime.datetime.strptime(item_pk[1], "%Y-%m-%d").date()
    table = EmployeeComplete.objects.filter(
        empid=item_pk[0],
        dat__year=date.year,
        dat__month=date.month,
        dat__day=date.day
    )
    if len(table) > 0:
        present = True
    return JsonResponse({"response":present})

@login_required
def getDocDetails(request, graph_id):
    print(graph_id)
    docDetails = Document.objects.get(file_title=graph_id.split("_")[1])
    start = docDetails.file_from
    end = docDetails.file_to
    delta = end - start
    dates, count = [], []
    for i in range(delta.days + 1):
        date = (start + datetime.timedelta(days=i))
        dates.append(date)
        temp = Employee.objects.filter(attendence_date = date).exclude(work_time="NS").extra({'date':"attendence_date"}).values('date').annotate(date_count=Count('attendence_date'))
        count.append(temp.values_list('date_count', flat=True)[0])
    data = {
        'count': count,
        'labels': dates
    }
    return JsonResponse(data)

@login_required
def getQuote(request):
    from random import randint
    quoteNums = randint(15, 25)
    print(quoteNums)
    q,a = [],[]
    for i in range(quoteNums):
        q.append("<p>"+quotes[randint(0,len(quotes))]["quoteText"]+"</p>")
        a.append(quotes[randint(0,len(quotes))]["quoteAuthor"])
    data = {
        "quotes": q,
        "authors": a
    }
    return JsonResponse(data)

@csrf_exempt
@login_required
def uploadMaster(request):
    if request.method == 'POST' and request.FILES['input-b6[]']:
        fileXL = request.FILES['input-b6[]']
        masterExcel = os.path.join(settings.STATICFILES_DIRS[0],"masters")
        if not os.path.exists('%s/archive/'%masterExcel):
            os.mkdir('%s/archive/'%masterExcel)
        for filename in os.listdir(masterExcel):
            if not filename.startswith("archive"):
                os.renames(os.path.join(masterExcel,filename), os.path.join(masterExcel,"archive/%s"%filename))
        with open(os.path.join(masterExcel, str(fileXL)), 'wb+') as destination:
            for chunk in fileXL.chunks():
                destination.write(chunk)
        return JsonResponse({})
    else:
        return JsonResponse({"response": "Unauthorized"})

def alive(request):
    return HttpResponse("Server alive --- to avoid server timeout")

def mutateModel():
    Employee.add_to_class('sample-field', models.CharField(max_length=255, blank=True))
    os.system("python manage.py makemigrations")
    os.system("python manage.py migrate")
    with open("./sample.py", 'w') as sampleFile:
        sampleFile.write("print('new field added')")
        sampleFile.flush()
        sampleFile.close()

def validateData(request):
    print("Validating Data")
    masterDir = os.path.join(settings.STATICFILES_DIRS[0],"masters")
    for fileNames in os.listdir(masterDir):
        if not fileNames.startswith("archive"):
            masterFile = os.path.join(masterDir,fileNames)
            break
    missing = []
    docs = Document.objects.filter(processed=False)
    wb = openpyxl.load_workbook(masterFile)
    employee_sheet = wb.get_sheet_names()[1]
    worksheet = wb.get_sheet_by_name(employee_sheet)
    empMaster = [worksheet['{}{}'.format('A',row)].value for row in range(2,worksheet.max_row+1)]
    for doc in docs:
        dataFilePath = os.path.join(settings.MEDIA_ROOT,doc.file_title)
        wb = openpyxl.load_workbook(dataFilePath)
        employee_sheet = wb.get_sheet_names()[0]
        worksheet = wb.get_sheet_by_name(employee_sheet)
        empIds = []
        for row in range(6,worksheet.max_row+1):
            temp = worksheet['{}{}'.format('E',row)].value
            if temp not in empIds:
                empIds.append(temp)
        tempMissing = [no for no in empIds if no not in empMaster]
        print(tempMissing)
        for item in tempMissing:
            if any(str(item).startswith(a) for a in ['350', 'INT', '550']):
                missing.append(str(item))
    if len(missing) > 0:
        print(missing)
        sendFailedMail(missing, request.user.email)
    return missing

def bad_req(request, **kwargs):
    return render(request, '400.html', status=400)
def perm_denied(request, **kwargs):
    return render(request, '403.html', status=403)
def page_not_found(request, **kwargs):
    return render(request, '404.html', status=404)
def server_error(request, **kwargs):
    return render(request, '500.html', status=500)
