import datetime
import nested_dict
import openpyxl
import time
import json
import os
import gc
import configparser
import pickle
import sys
from confiParse import ruleParser

def normaliseDict(items):
    for keys,values in items.items():
        if isinstance(values,datetime.datetime):
            items[keys] = values.isoformat()

        if not isinstance(values, datetime.datetime) and not isinstance(values, str):
            normaliseDict(values)
    return items

def findNextKey(dictionary,empid,key):
    i = 0
    for keys in dictionary[empid].keys():
        if str(key) == str(keys):
            i += 1
        elif i == 1:
            return str(keys)

def merge(a, b, path=None):
    if path is None: path = []
    for key in b:
        if key in a:
            if isinstance(a[key], dict) and isinstance(b[key], dict):
                merge(a[key], b[key], path + [str(key)])
            elif a[key] == b[key]:
                pass # same leaf value
            else:
                raise Exception('Conflict at %s' % '.'.join(path + [str(key)]))
        else:
            a[key] = b[key]
    return a

def parseExcel(excelsheet):
    entryList =nested_dict.nested_dict()
    gateList = nested_dict.nested_dict()
    employee = nested_dict.nested_dict()
    print("Reading {}...".format(excelsheet))
    block1 = openpyxl.load_workbook("./Dummy Samples/{}".format(excelsheet),read_only = True)
    activeSheet = block1.get_sheet_by_name(block1.get_sheet_names()[0])
    masterDate = ""
    print("Time to load cells from XL using openpyxl")
    t = time.time()
    cells = [rows for rows in activeSheet.iter_rows(row_offset=5)]
    print(time.time()-t)
    for cell in cells:
        if any(str((cell[4].value)).startswith(substring) for substring in ["350","550","INT"]):
            # print(str((cell[4].value)))
            if cell[0].value == None:
                break
            # print(cell[0].value,cell[1].value,str((cell[4].value)),str(cell[9].value).lower(),str(cell[10].value).lower())
            dateEmp, timeEmp, empid, gate, inout =cell[0].value,cell[1].value,str((cell[4].value)),str(cell[9].value).lower(),str(cell[10].value).lower()
            if str(gate) in primaryEntryPoints:
                if gateList[gate] == {}:
                    gateList[gate] = 0

                try:
                    cellDateTime = datetime.datetime.strptime("{} {}".format(dateEmp,timeEmp), "%d/%m/%Y %H:%M:%S")
                except:
                    cellDateTime = datetime.datetime.strptime("{} {}".format(dateEmp,timeEmp), "%d-%m-%Y %H:%M:%S")

                if cellDateTime.date() != masterDate:
                    masterDate = cellDateTime.date()
                    for keys in gateList.keys():
                        gateList[keys] +=1

                logEntry = "{} {}".format(gate,gateList[gate])

                if employee[empid][str(cellDateTime.date())][gate][logEntry] == {}:
                    entryList[gate][logEntry] = logEntry
                    employee[empid][str(cellDateTime.date())][gate][logEntry][inout] = (cellDateTime)
                else:
                    gateList[gate] += 1

                    if str(inout) in [key for key in employee[empid][str(cellDateTime.date())][gate][logEntry].keys()]:
                        logEntry = "{} {}".format(gate,gateList[gate])
                        employee[empid][str(cellDateTime.date())][gate][logEntry][inout] = (cellDateTime)
                        entryList[gate][logEntry] = logEntry
                    else:
                        employee[empid][str(cellDateTime.date())][gate][logEntry][inout] = (cellDateTime)
                        del entryList[gate][logEntry]

    return(employee)

gc.enable()
now = time.time()
i = 1
files = os.listdir("./Dummy Samples")
files = [files for files in files if not str(files).startswith("~$")]
masterEmployee = nested_dict.nested_dict()
primaryEntryPoints = ['biometric door','reception','main door 2','main door 1']
finalReport = openpyxl.Workbook()
finalSheet = finalReport.active
finalSheet.title = 'Master'
reprocess = False
config = configparser.ConfigParser()
config.read("./config.ini")

for xlfile in files:
    sizeF = os.path.getsize("./Dummy Samples/%s"%xlfile)
    try:
        prevSize = config.get("size",str(xlfile))
        if int(prevSize) != int(os.path.getsize("./Dummy Samples/%s"%xlfile)):
            reprocess = True
            print("File changed.. Reprocesing...")
            config.set('size', str(xlfile), str(sizeF))
            config.write(open("config.ini","w"))
    except Exception as e:
        print(e,"\nNo config item found, creating new item...")
        config.set('size', str(xlfile), str(sizeF))
        config.write(open("config.ini","w"))
        reprocess = True

    if reprocess:
        if not str(xlfile).startswith("~$"):
            empDictTemp = parseExcel(xlfile)
            merge(masterEmployee,empDictTemp)
            del empDictTemp
    gc.collect()
dumpableDict = normaliseDict(masterEmployee)

if not os.path.exists("./masterEmployee.json"):
    print("Json file missing, reprocessing the data..")
    reprocess = True

if reprocess:
    with open("masterEmployee.json","w") as jsonFile:
        (json.dump(dumpableDict,jsonFile,indent=3))
else:
    print("No file Change detected, loading data from json")
    with open("masterEmployee.json","r") as jsonFile:
        masterEmployee =json.load(jsonFile)
        masterEmployee =  nested_dict.nested_dict(masterEmployee)

# print("After Excel",memory()/(1024*1024*8)," MB")
# with open("masterEmployee.json","w") as jsonFile:
#     (json.dump(dumpableDict,jsonFile,indent=3))
# #
# # for empid in masterEmployee.keys():
# #     for date in masterEmployee[empid].keys():
# #         totalOfficeTime = datetime.timedelta(0)
# #         for gate in masterEmployee[empid][date].keys():
# #             inFlag, outFlag = False, False
# #             for logEntry in masterEmployee[empid][date][gate].keys():
# #                 if masterEmployee[empid][date][gate][logEntry]['in'] == {}:
# #                     messIn = (int(empid),str(date),gate,"In missing")
# #                     inFlag = True
# #                     finalSheet.append(messIn)
# #
# #                 if masterEmployee[empid][date][gate][logEntry]['out'] == {}:
# #                     messOut = (int(empid),str(date),gate,"Out missing")
# #                     outFlag = True
# #                     finalSheet.append(messOut)
# #
# #                 if not inFlag and not outFlag:
# #                     timeDifference = abs(masterEmployee[empid][date][gate][logEntry]['out']-masterEmployee[empid][date][gate][logEntry]['in'])
# #                     totalOfficeTime += timeDifference
# #
# #         messTotal = (int(empid),str(date),totalOfficeTime)
# #         finalSheet.append(messTotal)
# #
# print("Before Diff",memory()/(1024*1024*8)," MB")
master = openpyxl.load_workbook("Shift Master Employee Master for Automation 27092017.xlsx",read_only = True)
activeShiftSheet = master.get_sheet_by_name("Shift Master")
activeEmployeeSheet = master.get_sheet_by_name("Employee Master")
shiftDict = nested_dict.nested_dict()
irregularShifts = []

employeeShift = {str(rows[0].value):[rows[i].value for i in range(1,len(rows))] for rows in activeEmployeeSheet.iter_rows(row_offset=1)}
for rows in activeShiftSheet.iter_rows(row_offset=3):
    timeString = str(rows[2].value).replace(".",":").replace(" ","").split("to")
    timeFormat = "%I:%M%p"
    timeConverted = [str(datetime.datetime.strptime(temp, timeFormat).time()) for temp in timeString]
    if shiftDict[rows[0].value] == {}:
        i = 0
        shiftDict[rows[0].value][(rows[0].value+"{}".format(i))] = timeConverted
        i += 1
    else:
        shiftDict[rows[0].value][(rows[0].value+"{}".format(i))] = timeConverted
        i +=1

for codes,entries in shiftDict.items():
    for entry in entries.keys():
        if datetime.datetime.strptime(shiftDict[codes][entry][1],"%H:%M:%S") - datetime.datetime.strptime(shiftDict[codes][entry][0],"%H:%M:%S") < datetime.timedelta(0):
            irregularShifts.append(codes)
fromDate = input("Enter the date range in YYYY/MM/DD-YYYY/MM/DD format: ")
fromDate = fromDate.split("-")
fromDate = [fromDate[0].split("/"),fromDate[1].split("/")]
print(fromDate)
d1 = datetime.date(int(fromDate[0][0]), int(fromDate[0][1]), int(fromDate[0][2]))  # start date
d2 = datetime.date(int(fromDate[1][0]), int(fromDate[1][1]), int(fromDate[1][2]))  # end date
delta = d2 - d1         # timedelta
datesRange = []
datesRange.append("Emp ID")
for i in range(delta.days + 1):
    d = str(d1 + datetime.timedelta(days=i))
    datesRange.append(d)
finalSheet.append(datesRange)

for empid,dates in masterEmployee.items():
    mess = {}
    # mess[empid] = empid
    shiftCode = employeeShift[empid][4]
    for dateR in datesRange:
        if dateR not in list(masterEmployee[empid].keys()) and dateR != "Emp ID":
            mess[dateR] = "NS"
    for date,gates in (dates.items()):
        if shiftCode not in irregularShifts:
            inFlag,outFlag = False, False
            topGate = list(gates.keys())[0]
            bottomGate = list(gates.keys())[len(gates.keys())-1]
            topEntry = list(list(gates.values())[0].keys())[0]
            bottomEntry = list(list(gates.values())[0].keys())[len(list(gates.values())[0].keys())-1]
            inVal = masterEmployee[empid][date][topGate][topEntry]['in']
            outVal = masterEmployee[empid][date][bottomGate][bottomEntry]['out']
            # print(empid,"<>",topGate,"<>",bottomGate,"<>",inVal,"<>",outVal)
            if inVal == {}:
                inFlag= True
            if outVal == {}:
                outFlag = True
            if not inFlag and not outFlag:
                timeDifference = abs(datetime.datetime.strptime(outVal, "%Y-%m-%dT%H:%M:%S")-datetime.datetime.strptime(inVal, "%Y-%m-%dT%H:%M:%S"))
            if  inFlag or outFlag:
                mess[date] = "SIOM"
            else:
                mess[date] = timeDifference
                # code = ruleParser(timeDifference,None,None)
                # mess[date] += code
        else:
            nextDate = findNextKey(masterEmployee,empid,date)
            if nextDate != None:
                dateCond = (datetime.datetime.strptime(nextDate, "%Y-%m-%d") - datetime.timedelta(days=1))
                if date != str(dateCond.date()):
                    mess[date] = "SIOM"
                else:
                    mergedTemp = merge(masterEmployee[empid][date],masterEmployee[empid][nextDate])
                    # print("Created temp dict")
                    dateShiftedDict = nested_dict.nested_dict()
                    updated = False
                    actualshiftin = [datetime.datetime.strptime(shiftDict[shiftCode][list(shiftDict[shiftCode].keys())[i]][0], "%H:%M:%S").time() for i in range(len(list(shiftDict[shiftCode].keys())))]
                    actualshiftout = [datetime.datetime.strptime(shiftDict[shiftCode][list(shiftDict[shiftCode].keys())[i]][1], "%H:%M:%S").time() for i in range(len(list(shiftDict[shiftCode].keys())))]
                    inFlag,outFlag = False, False
                    for gate in mergedTemp.keys():
                        for entries in mergedTemp[gate].keys():
                            if mergedTemp[gate][entries]["in"] == {} or mergedTemp[gate][entries]["out"] == {}:
                                pass
                            else:
                                inVal = datetime.datetime.strptime(mergedTemp[gate][entries]["in"], "%Y-%m-%dT%H:%M:%S").time()
                                outVal = datetime.datetime.strptime(mergedTemp[gate][entries]["out"], "%Y-%m-%dT%H:%M:%S").time()
                                if any(inVal > ain for ain in actualshiftin) and any(outVal < aout for aout in actualshiftout) :
                                    updated = True
                                    dateShiftedDict[entries] = mergedTemp[gate][entries]
                    del mergedTemp
                    if updated:
                        inVal = dateShiftedDict[list(dateShiftedDict.keys())[0]]["in"]
                        outVal = dateShiftedDict[list(dateShiftedDict.keys())[len(dateShiftedDict.keys())-1]]["out"]
                        inFlag,outFlag = False, False
                        if inVal == {}:
                            inFlag= True
                        if outVal == {}:
                            outFlag = True
                        if not inFlag and not outFlag:
                            timeDifference = abs(datetime.datetime.strptime(outVal, "%Y-%m-%dT%H:%M:%S")-datetime.datetime.strptime(inVal, "%Y-%m-%dT%H:%M:%S"))
                        if  inFlag or outFlag:
                            mess[date] = "SIOM"
                        else:
                            mess[date] = timeDifference
                            # code = ruleParser(timeDifference,actualshiftin,actualshiftout)
                            # mess[date] += code
                    else:
                        mess[date] = "NS"
            else:
                mess[date] = "NS"
    sort = {k:mess[k] for k in sorted(list(mess.keys()))}
    del mess
    sort = list(sort.values())
    sort.insert(0,empid)
    finalSheet.append(sort)
    del sort

finalReport.save("finalReport.xlsx")
# print("After Diff",memory()/(1024*1024*8)," MB")
print(time.time() - now)
