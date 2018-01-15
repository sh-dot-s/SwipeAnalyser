from django.conf import settings
import datetime
import nested_dict
import openpyxl
import time
import json
import os
import gc
import configparser
import pickle
import sys

dateRangeDict = nested_dict.nested_dict()
config = configparser.ConfigParser()
configPath = os.path.join(settings.BASE_DIR,"config.ini")
config.read(configPath)
gracetimein = datetime.timedelta(hours=int(config.get("other","gracetimein")))
gracetimeout = datetime.timedelta(hours=int(config.get("other","gracetimeout")))

def normaliseDict(items):
    for keys,values in items.items():
        if isinstance(values,datetime.datetime):
            items[keys] = values.isoformat()

        if not isinstance(values, datetime.datetime) and not isinstance(values, str):
            normaliseDict(values)
    return items

def findNextKey(dictionary,empid,key):
    i = 0
    for keys in dictionary[empid].keys():
        if str(key) == str(keys):
            i += 1
        elif i == 1:
            return str(keys)

def merge(a, b, path=None):
    if path is None: path = []
    for key in b:
        if key in a:
            if isinstance(a[key], dict) and isinstance(b[key], dict):
                merge(a[key], b[key], path + [str(key)])
            elif a[key] == b[key]:
                pass # same leaf value
            else:
                raise Exception('Conflict at %s' % '.'.join(path + [str(key)]))
        else:
            a[key] = b[key]
    return a

def dateTimeConv(obj):
    return(datetime.datetime.strptime(obj, "%Y-%m-%dT%H:%M:%S"))

def dateConv(obj):
    return(datetime.datetime.strptime(obj, "%Y-%m-%d"))

def timeConv(obj):
    # print(datetime.datetime.strptime(obj, "%H:%M:%S").time())
    return(datetime.datetime.strptime(obj, "%H:%M:%S").time())

def nextDate(obj):
    return(str((datetime.datetime.strptime(obj, "%Y-%m-%d")+datetime.timedelta(days=1)).date()))

def parseExcel(excelsheet, primaryEntryPoints):
    global dateRangeDict
    dateRangeDict = nested_dict.nested_dict()
    entryList =nested_dict.nested_dict()
    gateList = nested_dict.nested_dict()
    employee = nested_dict.nested_dict()
    print("Reading {}...".format(excelsheet))
    block1 = openpyxl.load_workbook(excelsheet,read_only = True)
    activeSheet = block1.get_sheet_by_name(block1.get_sheet_names()[0])
    masterDate = ""
    print("Time to load cells from XL using openpyxl")
    t = time.time()
    cells = [rows for rows in activeSheet.iter_rows(row_offset=5)]
    print(time.time()-t)
    for cell in cells:
        if any(str((cell[4].value)).startswith(substring) for substring in ["350","550","INT"]):
            # print(str((cell[4].value)))
            if cell[0].value == None:
                break
            # print(cell[0].value,cell[1].value,str((cell[4].value)),str(cell[9].value).lower(),str(cell[10].value).lower())
            dateEmp, timeEmp, empid, gate, inout =cell[0].value,cell[1].value,str((cell[4].value)),str(cell[9].value).lower(),str(cell[10].value).lower()

            try:
                if isinstance(dateEmp, datetime.datetime):
                    formattedDate = dateEmp.date()
                else:
                    formattedDate = datetime.datetime.strptime(dateEmp, "%d/%m/%Y").date()
            except:
                if isinstance(dateEmp, datetime.datetime):
                    formattedDate = dateEmp.date()
                else:
                    formattedDate = datetime.datetime.strptime(dateEmp, "%d-%m-%Y").date()
            if dateRangeDict[formattedDate] == {}:
                dateRangeDict[formattedDate] = formattedDate
            else:
                pass

            if str(gate) in primaryEntryPoints:
                if gateList[gate] == {}:
                    gateList[gate] = 0

                try:
                    cellDateTime = datetime.datetime.strptime("{} {}".format(dateEmp,timeEmp), "%d/%m/%Y %H:%M:%S")
                except:
                    cellDateTime = datetime.datetime.strptime("{} {}".format(dateEmp,timeEmp), "%d-%m-%Y %H:%M:%S")

                if cellDateTime.date() != masterDate:
                    masterDate = cellDateTime.date()
                    for keys in gateList.keys():
                        gateList[keys] +=1

                logEntry = "{} {}".format(gate,gateList[gate])

                if employee[empid][str(cellDateTime.date())][gate][logEntry] == {}:
                    entryList[gate][logEntry] = logEntry
                    employee[empid][str(cellDateTime.date())][gate][logEntry][inout] = (cellDateTime)
                else:
                    gateList[gate] += 1

                    if str(inout) in [key for key in employee[empid][str(cellDateTime.date())][gate][logEntry].keys()]:
                        logEntry = "{} {}".format(gate,gateList[gate])
                        employee[empid][str(cellDateTime.date())][gate][logEntry][inout] = (cellDateTime)
                        entryList[gate][logEntry] = logEntry
                    else:
                        employee[empid][str(cellDateTime.date())][gate][logEntry][inout] = (cellDateTime)
                        del entryList[gate][logEntry]

    return(employee, cells)

def load_master_emp(datafile):
    global config, dateRangeDict
    gc.enable()
    now = time.time()
    i = 1
    files = datafile
    files = [files for files in files if not str(files).startswith("~$")]
    masterEmployee = nested_dict.nested_dict()
    primaryEntryPoints = ['biometric door','reception','reception 1','reception 2','main door 2','main door 1','b-1a main door 2','b-1a main door 1','biometric entry','b-1a odc front','b-1a odc exit','odc front','odc exit','b-1a reception 2','b-1a reception 1']
    reprocess = False
    dataFilePath = os.path.join(settings.MEDIA_ROOT,datafile)
    sizeF = os.path.getsize(dataFilePath)
    try:
        prevSize = config.get("size",str(datafile))
        print(prevSize,dataFilePath,datafile)
        # if int(prevSize) == int(os.path.getsize(dataFilePath)):
        reprocess = True
        print("File changed.. Reprocesing...")
        config.set('size', str(datafile), str(sizeF))
        config.write(open(configPath,"w"))
    except Exception as e:
        print(e,"\nNo config item found, creating new item...")
        config.set('size', str(datafile), str(sizeF))
        config.write(open(configPath,"w"))
        reprocess = True

    if reprocess:
        if not str(datafile).startswith("~$"):
            empDictTemp, cells = parseExcel(dataFilePath, primaryEntryPoints)
            merge(masterEmployee,empDictTemp)
            del empDictTemp
    gc.collect()
    try:
        d1, d2 = config.get("other","fromdate{}".format(datafile)),config.get("other","todate{}".format(datafile))
        d1 = datetime.datetime.strptime(d1, "%Y-%m-%d").date()
        d2 = datetime.datetime.strptime(d2, "%Y-%m-%d").date()

    except:
        dateRangeDict = [k for k in sorted(list(dateRangeDict))]
        d1, d2 = dateRangeDict[0], dateRangeDict[len(dateRangeDict)-1]
        config.set("other","fromdate{}".format(datafile),str(d1))
        config.set("other","todate{}".format(datafile),str(d2))
        config.write(open("config.ini","w"))

    dumpableDict = normaliseDict(masterEmployee)
    if not os.path.exists("./masterEmployee{}_{}.json".format(d1,d2)):
        print("Json file missing, reprocessing the data..")
        reprocess = True

    if reprocess:
        with open("masterEmployee{}_{}.json".format(d1,d2),"w") as jsonFile:
            (json.dump(dumpableDict,jsonFile,indent=3))
    else:
        print("No file Change detected, loading data from json")
        with open("masterEmployee{}_{}.json".format(d1,d2),"r") as jsonFile:
            masterEmployee =json.load(jsonFile)
            masterEmployee =  nested_dict.nested_dict(masterEmployee)
    return(masterEmployee, cells)

def shift_emp_masters():
    masterDir = os.path.join(settings.STATICFILES_DIRS[0],"masters")
    masterFile = os.path.join(masterDir,os.listdir(masterDir)[0])
    master = openpyxl.load_workbook(masterFile,read_only = True)
    activeShiftSheet = master.get_sheet_by_name("Shift Master")
    activeEmployeeSheet = master.get_sheet_by_name("Employee Master")
    shiftDict = nested_dict.nested_dict()
    irregularShifts = []
    employeeShift = {str(rows[0].value):[rows[i].value for i in range(1,len(rows))] for rows in activeEmployeeSheet.iter_rows(row_offset=1)}
    for rows in activeShiftSheet.iter_rows(row_offset=3):
        timeString = str(rows[2].value).replace(".",":").replace(" ","").split("to")
        timeFormat = "%I:%M%p"
        timeConverted = [str(datetime.datetime.strptime(temp, timeFormat).time()) for temp in timeString]
        if shiftDict[rows[0].value] == {}:
            i = 0
            shiftDict[rows[0].value][(rows[0].value+"{}".format(i))] = timeConverted
            i += 1
        else:
            shiftDict[rows[0].value][(rows[0].value+"{}".format(i))] = timeConverted
            i +=1

    for codes,entries in shiftDict.items():
        for entry in entries.keys():
            if datetime.datetime.strptime(shiftDict[codes][entry][1],"%H:%M:%S") - datetime.datetime.strptime(shiftDict[codes][entry][0],"%H:%M:%S") < datetime.timedelta(0):
                irregularShifts.append(codes)
    # print(shiftDict,employeeShift,irregularShifts)
    return (shiftDict,employeeShift,irregularShifts)

def calc_time(masterEmployee,employeeShift,irregularShifts,shiftDict,fileName):
    global dateRangeDict
    finalReport = openpyxl.Workbook()
    finalSheet = finalReport.active
    finalSheet.title = 'Master'
    # print(dateRangeDict)
    if dateRangeDict == {}:
        d1, d2 = config.get("other","fromdate{}".format(fileName)),config.get("other","todate{}".format(fileName))
        d1 = datetime.datetime.strptime(d1, "%Y-%m-%d").date()
        d2 = datetime.datetime.strptime(d2, "%Y-%m-%d").date()

    else:
        dateRangeDict = [k for k in sorted(list(dateRangeDict))]
    # d1 = datetime.date(int(fromDate[0][0]), int(fromDate[0][1]), int(fromDate[0][2]))  # start date
    # d2 = datetime.date(int(fromDate[1][0]), int(fromDate[1][1]), int(fromDate[1][2]))  # end date
        d1, d2 = dateRangeDict[0], dateRangeDict[len(dateRangeDict)-1]
        config.set("other","fromdate{}".format(fileName),str(d1))
        config.set("other","todate{}".format(fileName),str(d2))
        config.write(open("config.ini","w"))

    delta = d2 - d1         # timedelta
    datesRange = []
    masterList = []
    primaryEntryPoints = ['biometric door','reception','reception 1','reception 2','main door 2','main door 1','b-1a main door 2','b-1a main door 1','biometric entry','b-1a odc front','b-1a odc exit','odc front','odc exit','b-1a reception 2','b-1a reception 1']
    datesRange.append("Emp ID")
    for i in range(delta.days + 1):
        d = str(d1 + datetime.timedelta(days=i))
        datesRange.append(d)
    finalSheet.append(datesRange)
    # masterList.append(datesRange)

    inoutDict = nested_dict.nested_dict()
    for keys, values in (masterEmployee.items_flat()):
        shiftCode = employeeShift[keys[0]][len(employeeShift[keys[0]])-1]
        if keys[4] == "in":
            if inoutDict[keys[0]][keys[1]]["firstin"] == {}:
                inoutDict[keys[0]][keys[1]]["firstin"] = values
            elif dateTimeConv(inoutDict[keys[0]][keys[1]]["firstin"]) >  dateTimeConv(values):
                inoutDict[keys[0]][keys[1]]["firstin"] = values
            if inoutDict[keys[0]][keys[1]]["lastin"] == {}:
                inoutDict[keys[0]][keys[1]]["lastin"] = values
            elif dateTimeConv(inoutDict[keys[0]][keys[1]]["lastin"]) < dateTimeConv(values):
                inoutDict[keys[0]][keys[1]]["lastin"] = values


        if keys[4] == "out":
            if inoutDict[keys[0]][keys[1]]["firstout"] == {}:
                inoutDict[keys[0]][keys[1]]["firstout"] = values
            elif dateTimeConv(inoutDict[keys[0]][keys[1]]["firstout"]) >  dateTimeConv(values):
                inoutDict[keys[0]][keys[1]]["firstout"] = values
            if inoutDict[keys[0]][keys[1]]["lastout"] == {}:
                inoutDict[keys[0]][keys[1]]["lastout"] = values
            elif dateTimeConv(inoutDict[keys[0]][keys[1]]["lastout"]) < dateTimeConv(values):
                inoutDict[keys[0]][keys[1]]["lastout"] = values

    for empid in masterEmployee.keys():
        shiftCode = employeeShift[empid][len(employeeShift[empid])-1]
        if shiftCode in ['C','D','E']:
            actualshiftin = [(datetime.datetime.strptime(shiftDict[shiftCode][list(shiftDict[shiftCode].keys())[i]][0], "%H:%M:%S")-gracetimein).time() for i in range(len(list(shiftDict[shiftCode].keys())))]
            actualshiftout = [(datetime.datetime.strptime(shiftDict[shiftCode][list(shiftDict[shiftCode].keys())[i]][1], "%H:%M:%S")).time() for i in range(len(list(shiftDict[shiftCode].keys())))]
            actualshiftoutgrace = [(datetime.datetime.strptime(shiftDict[shiftCode][list(shiftDict[shiftCode].keys())[i]][1], "%H:%M:%S")+gracetimeout).time() for i in range(len(list(shiftDict[shiftCode].keys())))]
            for keys,values in masterEmployee[empid].items_flat():
                # if empid == "350040":
                #     print("INFO:",keys,values)
                if any(dateTimeConv(values).time() > ain for ain in actualshiftin):
                    if inoutDict[empid][keys[0]]["shiftin"] == {}:
                        inoutDict[empid][keys[0]]["shiftin"] = values
                    elif dateTimeConv(inoutDict[empid][keys[0]]["shiftin"]) > dateTimeConv(values):
                        inoutDict[empid][keys[0]]["shiftin"] = values
                if any(dateTimeConv(values).time() < aout for aout in actualshiftout) or any(dateTimeConv(values).time() < aout for aout in actualshiftoutgrace):
                    if inoutDict[empid][keys[0]]["shiftout"] == {}:
                        inoutDict[empid][keys[0]]["shiftout"] = values
                        # if empid == "350040":
                        #     print("FIRST:",keys,values)
                    elif dateTimeConv(inoutDict[empid][keys[0]]["shiftout"]) < dateTimeConv(values):
                        inoutDict[empid][keys[0]]["shiftout"] = values
                        # if empid == "350040":
                        #     print("UPDATE:",keys,values)

    timeDict = nested_dict.nested_dict()
    for empid in inoutDict.keys():
        shiftCode = employeeShift[empid][len(employeeShift[empid])-1]
        # predictedShift[empid] = predictShift(empid)

        for date in inoutDict[empid].keys():
            if shiftCode in ["D", "E"]:
                if inoutDict[empid][date]["lastout"] != {} or inoutDict[empid][date]["firstin"] != {}:
                    if not nextDate(date) in list(inoutDict[empid].keys()):
                        if inoutDict[empid][date]["shiftin"] == {}:
                            timeDict[empid][date] = "NS"
                        else:
                            timeDict[empid][date] = "SIOM"
                    else:
                        if inoutDict[empid][nextDate(date)]["shiftout"] != {} and inoutDict[empid][date]["shiftin"] != {}:
                            timeDiff = abs(dateTimeConv(inoutDict[empid][nextDate(date)]["shiftout"]) - dateTimeConv(inoutDict[empid][date]["shiftin"]))
                            timeDict[empid][date] = str(timeDiff)
                        else:
                            if inoutDict[empid][nextDate(date)]["shiftout"] == {} and inoutDict[empid][date]["shiftin"] == {}:
                                timeDict[empid][date] = "NS"
                            else:
                                timeDict[empid][date] = "SIOM"

                else:
                    timeDict[empid][date] = "SIOM"
            elif shiftCode == "C":
                updated = False
                if nextDate(date) in list(inoutDict[empid].keys()):
                    if inoutDict[empid][nextDate(date)]["firstin"] != {} and inoutDict[empid][date]["shiftin"] != {}:
                        if dateTimeConv(inoutDict[empid][nextDate(date)]["firstin"]).hour < 3:
                            timeDiffNext = dateTimeConv(inoutDict[empid][nextDate(date)]["firstin"]) - dateTimeConv(inoutDict[empid][date]["shiftin"])
                            timeDict[empid][date] = str(timeDiffNext)
                            updated = True
                            # print(timeDiffNext,empid)
                if inoutDict[empid][date]["lastout"] != {} and inoutDict[empid][date]["firstin"] != {}:
                    if updated:
                        pass
                    else:
                        timeDiff = dateTimeConv(inoutDict[empid][date]["lastout"]) - dateTimeConv(inoutDict[empid][date]["firstin"])
                        timeDict[empid][date] = str(timeDiff)
                        if inoutDict[empid][date]["shiftin"] != {} and inoutDict[empid][date]["shiftout"] != {}:
                            timeDiffShift = dateTimeConv(inoutDict[empid][date]["shiftout"]) - dateTimeConv(inoutDict[empid][date]["shiftin"])
                            timeDict[empid][date] += " "+str(timeDiffShift)
                else:
                    timeDict[empid][date] = "SIOM"
            else:
                if inoutDict[empid][date]["lastout"] != {} and inoutDict[empid][date]["firstin"] != {}:
                    timeDiff = dateTimeConv(inoutDict[empid][date]["lastout"]) - dateTimeConv(inoutDict[empid][date]["firstin"])
                    timeDict[empid][date] = str(timeDiff)
                else:
                    timeDict[empid][date] = "SIOM"

    # with open("ioFinal.json","w") as jsonFile:
    #     (json.dump(inoutDict,jsonFile,indent=3))
    # with open("ioTime.json","w") as jsonFile:
    #     (json.dump(timeDict,jsonFile,indent=3))
    return(timeDict, d1, d2)
